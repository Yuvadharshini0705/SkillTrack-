"""
routes/bulk_upload.py
Bulk Question Upload for SkillTrack
-------------------------------------
Single unified route for Excel + CSV uploads.
Course is selected by the admin before upload — no course_slug column needed.

EXPECTED QUESTIONS PER DAY:
  MCQ    → 10 questions
  Debug  →  4 questions
  Coding →  1 question
  Total  → 15 questions per day

Removed from template (auto-derived):
  xp_reward   → derived from difficulty via DIFF_XP_MAP
  time_limit  → derived from task_type via TYPE_TIME_MAP
  subtopic    → optional, omitted from template
  bug_count   → always 1
  is_recovery → always false at upload time
  tags        → auto-generated from topic + task_type

After a successful insert, calls notify_students_of_new_tasks() so students
whose current_day matches an uploaded day get a notification immediately.
"""

from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from db import get_db
from models import make_task, serialize, oid
from functools import wraps
import io

bulk_upload_bp = Blueprint("bulk_upload", __name__)


# ── Admin auth decorator ───────────────────────────────────────────────────────

def admin_required(f):
    @wraps(f)
    @jwt_required()
    def wrapped(*args, **kwargs):
        db   = get_db()
        user = db.users.find_one({"_id": oid(get_jwt_identity())})
        if not user or user.get("role") != "admin":
            return jsonify({"error": "Admin access required"}), 403
        return f(*args, **kwargs)
    return wrapped


# ── Constants ──────────────────────────────────────────────────────────────────

VALID_TASK_TYPES   = {"mcq", "debug", "coding"}
VALID_DIFFICULTIES = {"beginner", "intermediate", "advanced", "expert"}

# XP awarded on correct answer — derived from difficulty, not set per row
DIFF_XP_MAP = {
    "beginner":     10,
    "intermediate": 20,
    "advanced":     35,
    "expert":       50,
}

# difficulty_score used for leaderboard/skill weighting
DIFF_SCORE_MAP = {
    "beginner": 2, "intermediate": 5, "advanced": 8, "expert": 10,
}

# Default time limit per task type (seconds)
TYPE_TIME_MAP = {
    "mcq":    120,
    "debug":  300,
    "coding": 600,
}

# Expected questions per task type per day
EXPECTED_COUNTS = {
    "mcq":    10,
    "debug":  4,
    "coding": 1,
}

# Columns the admin must fill — deliberately minimal
REQUIRED_COLS = {
    "day", "task_type", "difficulty", "topic", "question", "correct_answer"
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _safe(val, default=""):
    try:
        import pandas as pd
        if pd.isna(val):
            return default
    except Exception:
        pass
    if val is None:
        return default
    return str(val).strip()


def _read_file(file_bytes, filename):
    fname = filename.lower()

    if fname.endswith((".xlsx", ".xls")):
        try:
            import pandas as pd
        except ImportError:
            return None, "pandas not installed. Run: pip install pandas openpyxl"
        try:
            df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
        except Exception as e:
            return None, f"Could not read Excel file: {e}"

    elif fname.endswith(".csv"):
        try:
            import pandas as pd
        except ImportError:
            return None, "pandas not installed. Run: pip install pandas"
        try:
            df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="utf-8-sig")
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="latin-1")
            except Exception as e:
                return None, f"Could not decode CSV file: {e}"
        except Exception as e:
            return None, f"Could not read CSV file: {e}"
    else:
        return None, "Only .xlsx, .xls, or .csv files are accepted"

    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    missing = REQUIRED_COLS - set(df.columns)
    if missing:
        return None, (
            f"Missing required columns: {sorted(missing)}. "
            "Required: day, task_type, difficulty, topic, question, correct_answer"
        )

    rows = df.where(df.notna(), other="").to_dict(orient="records")
    return rows, None


def _count_by_type_and_day(rows):
    """Return {day: {task_type: count}} for all rows."""
    summary = {}
    for row in rows:
        try:
            day = int(float(str(row.get("day", "1")) or "1"))
        except Exception:
            day = 0
        tt = str(row.get("task_type", "")).lower().strip()
        if day not in summary:
            summary[day] = {}
        summary[day][tt] = summary[day].get(tt, 0) + 1
    return summary


def _build_count_warnings(day_summary):
    """Non-blocking warnings when counts deviate from EXPECTED_COUNTS."""
    warnings = []
    for day, type_counts in sorted(day_summary.items()):
        for task_type, expected in EXPECTED_COUNTS.items():
            actual = type_counts.get(task_type, 0)
            if actual < expected:
                warnings.append(
                    f"Day {day}: {task_type.upper()} has {actual}/{expected} questions"
                )
            elif actual > expected:
                warnings.append(
                    f"Day {day}: {task_type.upper()} has {actual} questions "
                    f"(expected {expected}) — extras randomly chosen during tests"
                )
    return warnings


def _auto_tags(topic: str, task_type: str) -> list:
    """Generate tags automatically from topic and task_type."""
    base = topic.lower().replace(" ", "-")
    return [base, task_type, f"{base}-{task_type}"]


def _validate_and_build(row, row_num, course_id):
    errors = []

    task_type  = _safe(row.get("task_type", "")).lower()
    difficulty = _safe(row.get("difficulty", "")).lower()
    topic      = _safe(row.get("topic", ""))
    question   = _safe(row.get("question", ""))
    solution   = _safe(row.get("correct_answer", ""))

    if task_type not in VALID_TASK_TYPES:
        errors.append(f"task_type must be one of: {', '.join(sorted(VALID_TASK_TYPES))} (got '{task_type}')")
    if difficulty not in VALID_DIFFICULTIES:
        errors.append(f"difficulty must be one of: {', '.join(sorted(VALID_DIFFICULTIES))} (got '{difficulty}')")
    if not topic:
        errors.append("topic is required")
    if not question:
        errors.append("question is required")
    if not solution:
        errors.append("correct_answer is required")

    try:
        day = int(float(_safe(row.get("day", "1")) or "1"))
        if day < 1:
            raise ValueError
    except ValueError:
        errors.append(f"day must be a positive integer (got '{row.get('day')}')")
        day = 1

    if errors:
        return None, "; ".join(errors)

    # ── Auto-derive fields — not exposed in template ───────────────────────────
    xp_reward  = DIFF_XP_MAP.get(difficulty, 10)
    time_limit = TYPE_TIME_MAP.get(task_type, 300)
    tags       = _auto_tags(topic, task_type)

    # ── Optional columns — present in template but not required ───────────────
    title_raw   = _safe(row.get("title", ""))
    title       = title_raw or f"{topic} — {task_type.upper()}"
    explanation = _safe(row.get("explanation", ""))

    # ── Build content dict by task type ───────────────────────────────────────
    if task_type == "mcq":
        opt_a = _safe(row.get("option_a", ""))
        opt_b = _safe(row.get("option_b", ""))
        opt_c = _safe(row.get("option_c", ""))
        opt_d = _safe(row.get("option_d", ""))
        if not opt_a or not opt_b:
            return None, "MCQ requires at least option_a and option_b"
        correct_letter = solution.strip().upper()[:1]
        if correct_letter not in ("A", "B", "C", "D"):
            return None, f"correct_answer must be A/B/C/D for MCQ (got '{solution}')"
        opt_map = {"A": opt_a, "B": opt_b, "C": opt_c, "D": opt_d}
        options = [f"A) {opt_a}", f"B) {opt_b}"]
        if opt_c: options.append(f"C) {opt_c}")
        if opt_d: options.append(f"D) {opt_d}")
        content = {
            "question":       question,
            "options":        options,
            "correct_option": correct_letter,
            "correct_text":   opt_map.get(correct_letter, ""),
        }
        solution_final = correct_letter

    elif task_type == "debug":
        content = {
            "buggy_code":      question,
            "language":        _safe(row.get("language", "javascript")) or "javascript",
            "expected_output": _safe(row.get("expected_output", "")),
            "hints":           [h.strip() for h in _safe(row.get("hints", "")).split("|") if h.strip()],
            "bug_count":       1,   # always 1 — removed from template
        }
        solution_final = solution

    else:  # coding
        content = {
            "problem":      question,
            "starter_code": _safe(row.get("starter_code", "// write your solution here")),
            "language":     _safe(row.get("language", "javascript")) or "javascript",
            "test_cases":   [],
            "constraints":  [],
            "examples":     [],
        }
        solution_final = solution

    return dict(
        course_id        = course_id,
        title            = title,
        task_type        = task_type,
        difficulty       = difficulty,
        topic            = topic,
        subtopic         = "",          # not in template — left empty
        content          = content,
        solution         = solution_final,
        explanation      = explanation,
        difficulty_score = DIFF_SCORE_MAP.get(difficulty, 5),
        day_range_start  = day,
        day_range_end    = day,
        source           = "excel",
        xp_reward        = xp_reward,   # auto-derived, not from spreadsheet
        time_limit       = time_limit,  # auto-derived, not from spreadsheet
        tags             = tags,        # auto-generated
        is_recovery      = False,       # always false at upload time
    ), None


# =============================================================================
# ROUTE 1 — Bulk Upload
# POST /api/admin/bulk-upload
# =============================================================================

@bulk_upload_bp.route("/bulk-upload", methods=["POST"])
@admin_required
def bulk_upload():
    db  = get_db()
    uid = get_jwt_identity()

    course_id    = request.form.get("course_id", "").strip()
    auto_approve = request.form.get("auto_approve", "false").lower() in ("true", "1", "yes")
    preview_only = request.form.get("preview_only", "false").lower() in ("true", "1", "yes")

    if not course_id:
        return jsonify({"error": "course_id is required. Select a course before uploading."}), 400

    course = db.courses.find_one({"_id": oid(course_id)})
    if not course:
        return jsonify({"error": f"Course not found for id '{course_id}'"}), 404

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded. Send multipart/form-data with key 'file'"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Empty filename"}), 400

    file_bytes = file.read()
    rows, read_error = _read_file(file_bytes, file.filename)
    if read_error:
        return jsonify({"error": read_error}), 400

    if len(rows) == 0:
        return jsonify({"error": "The file has no data rows (only a header row found)"}), 400

    valid_docs   = []
    failed_rows  = []
    preview_rows = []

    for idx, row in enumerate(rows):
        row_num = idx + 2
        doc_kwargs, error = _validate_and_build(row, row_num, str(course["_id"]))

        if error:
            failed_rows.append({
                "row":    row_num,
                "title":  str(row.get("title", "") or row.get("topic", "") or f"Row {row_num}"),
                "reason": error,
            })
            preview_rows.append({
                "row":        row_num,
                "title":      str(row.get("title", "") or row.get("topic", "") or f"Row {row_num}"),
                "type":       str(row.get("task_type", "")),
                "difficulty": str(row.get("difficulty", "")),
                "day":        str(row.get("day", "")),
                "status":     "error",
                "reason":     error,
            })
        else:
            valid_docs.append(doc_kwargs)
            preview_rows.append({
                "row":        row_num,
                "title":      doc_kwargs["title"],
                "type":       doc_kwargs["task_type"],
                "difficulty": doc_kwargs["difficulty"],
                "day":        doc_kwargs["day_range_start"],
                "xp":         doc_kwargs["xp_reward"],
                "status":     "valid",
                "reason":     None,
            })

    all_rows_summary = _count_by_type_and_day([r for r in rows if r])
    count_warnings   = _build_count_warnings(all_rows_summary)

    # ── Preview mode ──────────────────────────────────────────────────────────
    if preview_only:
        return jsonify({
            "preview":          True,
            "course_name":      course.get("name", ""),
            "total_rows":       len(rows),
            "valid_count":      len(valid_docs),
            "error_count":      len(failed_rows),
            "rows":             preview_rows,
            "failed_rows":      failed_rows,
            "count_warnings":   count_warnings,
            "expected_per_day": EXPECTED_COUNTS,
            "day_summary":      all_rows_summary,
        }), 200

    # ── Insert valid rows ─────────────────────────────────────────────────────
    status   = "approved" if auto_approve else "pending"
    inserted = 0

    for kwargs in valid_docs:
        doc = make_task(
            status     = status,
            created_by = uid,
            **kwargs,
        )
        db.tasks.insert_one(doc)
        inserted += 1

    # ── Notify students if auto-approved ──────────────────────────────────────
    if inserted > 0 and auto_approve:
        try:
            from utils.assignment import notify_students_of_new_tasks
            notify_students_of_new_tasks(
                db,
                str(course["_id"]),
                course.get("name", ""),
            )
        except Exception as e:
            print(f"[UPLOAD] Notification error (non-fatal): {e}")

    return jsonify({
        "preview":          False,
        "course_name":      course.get("name", ""),
        "course_id":        str(course["_id"]),
        "total_rows":       len(rows),
        "inserted":         inserted,
        "error_count":      len(failed_rows),
        "auto_approved":    auto_approve,
        "status":           status,
        "failed_rows":      failed_rows[:100],
        "count_warnings":   count_warnings,
        "expected_per_day": EXPECTED_COUNTS,
        "message":          (
            f"{inserted} questions imported into '{course.get('name', '')}' "
            f"({'approved — students notified' if auto_approve else 'pending review'}). "
            f"{len(failed_rows)} rows skipped."
        ),
    }), 201


# =============================================================================
# ROUTE 2 — Download Template
# GET /api/admin/bulk-upload/template?format=xlsx  (default)
# GET /api/admin/bulk-upload/template?format=csv
#
# Minimal template: only columns the admin actually needs to fill.
# Removed: xp_reward, time_limit, subtopic, bug_count, is_recovery, tags
# Kept optional: title, explanation, language, starter_code, expected_output, hints
# =============================================================================

@bulk_upload_bp.route("/bulk-upload/template", methods=["GET"])
@admin_required
def download_template():
    fmt = request.args.get("format", "xlsx").lower()

    # ── Minimal headers — only what the admin needs ───────────────────────────
    HEADERS = [
        "day",           # required
        "task_type",     # required — mcq | debug | coding
        "difficulty",    # required — beginner | intermediate | advanced | expert
        "topic",         # required
        "title",         # optional — auto-generated if blank
        "question",      # required — MCQ: question text | debug: buggy code | coding: problem
        "option_a",      # MCQ only
        "option_b",      # MCQ only
        "option_c",      # MCQ only
        "option_d",      # MCQ only
        "correct_answer",# required — MCQ: A/B/C/D | debug/coding: solution text
        "explanation",   # optional — shown after student submits
        "language",      # debug/coding only — defaults to javascript
        "starter_code",  # coding only
        "expected_output",# debug only
        "hints",         # debug only — pipe-separated e.g. hint1|hint2
    ]

    # ── Example rows: 10 MCQ + 4 Debug + 1 Coding for Day 1 ──────────────────

    MCQ_TOPICS = [
        ("JavaScript Variables",
         "Which keyword declares a block-scoped variable that cannot be reassigned?",
         "var", "let", "const", "static", "C",
         "const declares a block-scoped, non-reassignable binding."),
        ("JavaScript Variables",
         "What is the scope of a var-declared variable inside a function?",
         "Block", "Module", "Function", "Global", "C",
         "var is function-scoped, not block-scoped."),
        ("JavaScript Functions",
         "Which syntax correctly defines an arrow function?",
         "function(){}", "=>function{}", "() => {}", "def()=>{}", "C",
         "Arrow functions use the => syntax."),
        ("JavaScript Arrays",
         "Which method adds an element to the END of an array?",
         "unshift", "push", "splice", "pop", "B",
         "push() appends to the end; unshift() prepends."),
        ("JavaScript Arrays",
         "What does Array.map() return?",
         "The same array", "A new array", "undefined", "A boolean", "B",
         "map() returns a new array without mutating the original."),
        ("JavaScript Objects",
         "How do you access a property 'name' on object obj?",
         "obj->name", "obj::name", "obj.name", "obj[name]", "C",
         "Dot notation (obj.name) or bracket notation (obj['name'])."),
        ("JavaScript Objects",
         "Which method returns an array of an object's own enumerable property names?",
         "Object.values()", "Object.keys()", "Object.entries()", "Object.props()", "B",
         "Object.keys() returns an array of property names."),
        ("JavaScript Loops",
         "Which loop is guaranteed to execute at least once?",
         "for", "while", "do...while", "for...of", "C",
         "do...while checks the condition after executing the body."),
        ("JavaScript Promises",
         "What does the await keyword do?",
         "Delays the code", "Pauses until promise resolves", "Creates a promise", "Rejects a promise", "B",
         "await pauses async function execution until the promise settles."),
        ("JavaScript DOM",
         "Which method selects the FIRST element matching a CSS selector?",
         "getElementById", "querySelector", "querySelectorAll", "getElementByClass", "B",
         "querySelector returns the first matching element."),
    ]

    DEBUG_TOPICS = [
        ("JavaScript Functions",
         "Debug: missing return",
         "const double = (n) => {\n  n * 2;\n};\nconsole.log(double(5));",
         "const double = (n) => n * 2;\nconsole.log(double(5));",
         "Missing return — arrow functions with {} need explicit return.",
         "Are you returning a value?|Arrow functions need explicit return with braces"),
        ("JavaScript Arrays",
         "Debug: wrong method",
         "const nums = [1, 2, 3];\nconst doubled = nums.forEach(n => n * 2);\nconsole.log(doubled);",
         "const nums = [1, 2, 3];\nconst doubled = nums.map(n => n * 2);\nconsole.log(doubled);",
         "forEach does not return a new array — use map() instead.",
         "Does forEach return anything?|Which array method returns a new array?"),
        ("JavaScript Objects",
         "Debug: undefined property",
         "const user = { name: 'Alice' };\nconsole.log(user.age.toString());",
         "const user = { name: 'Alice', age: 25 };\nconsole.log(user.age.toString());",
         "user.age is undefined — accessing .toString() on undefined throws.",
         "What happens when you access a missing property?|Check if the property exists first"),
        ("JavaScript Async",
         "Debug: missing await",
         "async function getData() {\n  const res = fetch('https://api.example.com/data');\n  return res.json();\n}",
         "async function getData() {\n  const res = await fetch('https://api.example.com/data');\n  return res.json();\n}",
         "fetch returns a Promise — you must await it before calling .json().",
         "Is fetch synchronous?|What does await do to a Promise?"),
    ]

    CODING = (
        "JavaScript Arrays",
        "Sum of Array",
        "Write a function sumArray(arr) that returns the sum of all numbers in the array.",
        "function sumArray(arr) { return arr.reduce((a, b) => a + b, 0); }",
        "Use Array.reduce() to accumulate the total.",
        "function sumArray(arr) {\n  // your code here\n}",
    )

    examples = []

    # 10 MCQ rows
    for i, (topic, question, a, b, c, d, answer, explanation) in enumerate(MCQ_TOPICS):
        examples.append([
            1, "mcq", "beginner", topic,
            f"{topic} Q{i+1}",
            question, a, b, c, d, answer, explanation,
            "", "", "", "",
        ])

    # 4 Debug rows
    for topic, title, buggy, solution, explanation, hints in DEBUG_TOPICS:
        examples.append([
            1, "debug", "beginner", topic,
            title,
            buggy, "", "", "", "", solution, explanation,
            "javascript", "", "", hints,
        ])

    # 1 Coding row
    topic, title, question, solution, explanation, starter = CODING
    examples.append([
        1, "coding", "beginner", topic,
        title,
        question, "", "", "", "", solution, explanation,
        "javascript", starter, "", "",
    ])

    # ── CSV export ────────────────────────────────────────────────────────────
    if fmt == "csv":
        import csv as csv_mod
        output = io.StringIO()
        writer = csv_mod.writer(output)
        writer.writerow(HEADERS)
        for ex in examples:
            writer.writerow(ex)
        output.seek(0)
        from flask import Response
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=skilltrack_questions_template.csv"},
        )

    # ── Excel export ──────────────────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Questions"

    # Column widths matching the 16 headers
    COL_WIDTHS = [6, 12, 14, 22, 28, 52, 16, 16, 16, 16, 16, 40, 12, 30, 20, 40]

    hdr_fill = PatternFill("solid", start_color="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin     = Side(style="thin", color="AAAAAA")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[1].height = 30

    MCQ_FILL    = PatternFill("solid", start_color="EBF5FB")
    DEBUG_FILL  = PatternFill("solid", start_color="FEF9E7")
    CODING_FILL = PatternFill("solid", start_color="FDEDEC")

    def row_fill(ex):
        t = str(ex[1]).lower()
        if t == "debug":  return DEBUG_FILL
        if t == "coding": return CODING_FILL
        return MCQ_FILL

    body_font = Font(name="Calibri", size=10)
    wrap      = Alignment(wrap_text=True, vertical="top")

    for r_offset, example in enumerate(examples, start=2):
        ws.row_dimensions[r_offset].height = 45
        fill = row_fill(example)
        for c_idx, val in enumerate(example, start=1):
            cell           = ws.cell(row=r_offset, column=c_idx, value=val)
            cell.fill      = fill
            cell.font      = body_font
            cell.alignment = wrap
            cell.border    = border

    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Dropdowns for task_type and difficulty
    dv_type = DataValidation(type="list", formula1='"mcq,debug,coding"', allow_blank=False)
    dv_diff = DataValidation(type="list", formula1='"beginner,intermediate,advanced,expert"', allow_blank=False)
    ws.add_data_validation(dv_type)
    ws.add_data_validation(dv_diff)
    dv_type.sqref = "B2:B2000"
    dv_diff.sqref = "C2:C2000"

    # ── Instructions sheet ────────────────────────────────────────────────────
    inst = wb.create_sheet("Instructions")
    inst.column_dimensions["A"].width = 24
    inst.column_dimensions["B"].width = 90

    NOTES = [
        ("COLUMN",              "DESCRIPTION / RULES"),
        ("WHAT'S AUTO-DERIVED", ""),
        ("xp_reward",           "Auto-set from difficulty: beginner=10 · intermediate=20 · advanced=35 · expert=50"),
        ("time_limit",          "Auto-set from task type: MCQ=2 min · Debug=5 min · Coding=10 min"),
        ("tags",                "Auto-generated from topic + task_type — no need to fill"),
        ("bug_count",           "Always 1 — removed from template"),
        ("is_recovery",         "Always false at upload time — admin can change later"),
        ("",                    ""),
        ("REQUIRED COLUMNS",    ""),
        ("day",                 "Day number (integer ≥ 1). Matches the student's current_day."),
        ("task_type",           "One of: mcq | debug | coding"),
        ("difficulty",          "One of: beginner | intermediate | advanced | expert"),
        ("topic",               "Broad topic label e.g. 'JavaScript Variables'"),
        ("question",            "MCQ: question text  ·  debug: buggy code  ·  coding: problem statement"),
        ("correct_answer",      "MCQ: A/B/C/D  ·  debug/coding: the correct solution or fixed code"),
        ("",                    ""),
        ("OPTIONAL COLUMNS",    ""),
        ("title",               "Short card title — auto-generated as 'Topic — TYPE' if blank"),
        ("explanation",         "Why the answer is correct, shown after the student submits"),
        ("language",            "debug/coding only — javascript | python | java | bash (default: javascript)"),
        ("starter_code",        "coding only — scaffold shown to student in the editor"),
        ("expected_output",     "debug only — what the fixed code should print or return"),
        ("hints",               "debug only — pipe-separated hints e.g. 'Check the return value|Is the method correct?'"),
        ("option_a – option_d", "MCQ only — the 4 answer choices (plain text, no letter prefix needed)"),
        ("",                    ""),
        ("QUESTION STRUCTURE",  ""),
        ("Per day target",      "MCQ: 10  ·  Debug: 4  ·  Coding: 1  (Total: 15 per day)"),
        ("Extra rows",          "Upload more than the target for the same day/type — extras are randomly chosen at test time"),
        ("",                    ""),
        ("SCHEDULING",          ""),
        ("How days work",
         "Students see tasks for their current_day. After submitting, their next day unlocks at IST midnight. "
         "Upload all 15 questions for Day 1, then Day 2, etc."),
    ]

    hdr_fill2    = PatternFill("solid", start_color="1E3A5F")
    hdr_font2    = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", start_color="2D6A9F")
    section_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    auto_fill    = PatternFill("solid", start_color="E8F5E9")
    auto_font    = Font(name="Calibri", size=10, bold=True, color="1B5E20")
    bold_font    = Font(name="Calibri", size=10, bold=True)
    plain_font   = Font(name="Calibri", size=10)
    wrap_align   = Alignment(wrap_text=True, vertical="top")

    SECTION_HEADERS = {"WHAT'S AUTO-DERIVED", "REQUIRED COLUMNS", "OPTIONAL COLUMNS",
                       "QUESTION STRUCTURE", "SCHEDULING"}

    for r_idx, (col, desc) in enumerate(NOTES, start=1):
        c1 = inst.cell(row=r_idx, column=1, value=col)
        c2 = inst.cell(row=r_idx, column=2, value=desc)
        if r_idx == 1:
            c1.font = hdr_font2; c2.font = hdr_font2
            c1.fill = hdr_fill2; c2.fill = hdr_fill2
        elif col in SECTION_HEADERS:
            c1.font = section_font; c2.font = section_font
            c1.fill = section_fill; c2.fill = section_fill
        elif col in ("xp_reward", "time_limit", "tags", "bug_count", "is_recovery"):
            c1.font = auto_font;  c2.font = plain_font
            c1.fill = auto_fill;  c2.fill = auto_fill
        elif col:
            c1.font = bold_font;  c2.font = plain_font
        else:
            c1.font = plain_font; c2.font = plain_font
        c1.alignment = wrap_align
        c2.alignment = wrap_align
        inst.row_dimensions[r_idx].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="skilltrack_questions_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )